import openpyxl

def main():
    TotalCount=1
    wb = openpyxl.load_workbook('~1k comments.xlsx')
    sheet = wb.active

    reqd_ids = []
    # while TotalCount < sheet.max_row:
    while TotalCount <= sheet.max_row:
        print(TotalCount)
        comment = sheet["A" + str(TotalCount)].value
        comment = comment.strip("',.\"")
        comment = comment.split()
        if '@' in comment[0]:
            print(comment[0])
            comment = comment[1:]
        comment = ' '.join(comment)
        # print(comment)
        sheet["A"+str(TotalCount)] = comment

        TotalCount += 1
        # if(len(comment.split()) <= 2):
        #     reqd_ids += [str(TotalCount)]
            # print("GHAPLA: " + TotalCount)

    wb.save('~1k comments.xlsx')
    # print(' '.join(reqd_ids))

main()