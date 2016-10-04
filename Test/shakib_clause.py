import sys
sys.path.append(r"C:\Users\Shabab\PycharmProjects\Thesis\Test")
import preprocessing_clause_level

import urllib
import urllib.request
import re
from bs4 import BeautifulSoup

TotalCount = 0
def generate_links(theurl,label=""):
    # preprocessing_clause_level.test()
    #theurl = a
    thepage = urllib.request.urlopen(theurl).read()
    # print(thepage)
    soup = BeautifulSoup(thepage, "html.parser")

    # print("length: " + str(len(soup('item'))))

    target = open("shakib_links.txt", 'a')
    # target.write("Label: " + label + '\n')
    try:
        # line1 = soup.find_all('div', {"class": "teams"}).find('a')
        links = soup.find_all('item')
        # links = soup.find_all('link', 'all')
        print(str(len(links)))
        for l in links:
            # print(l)
            link = l.find_all('link')
            for a_link in link[0]:
                print(a_link)
                target.write(str(a_link))
                target.write("\n")

    except Exception as e:
        print(str(e))
    # target.write('\n\n\n')

def generate_comments(theurl):
    thepage = urllib.request.urlopen(theurl).read()
    soup = BeautifulSoup(thepage, "html.parser")

    try:
        div = soup.find_all('div', 'comment-tabs-wrap')
        # links = soup.find_all('link', 'all')
        # print(str(len(div)))
        for d in div:
            comments = d.find_all('p')
            # dates = d.find_all('span', 'date')
            all_comments = set()
            # all_dates = []
            for a_comment in comments:
                # print(a_comment.text)
                comment = a_comment.text.strip()
                # date = a_date.text.strip()
                if comment=="No featured comments at the moment.":
                    continue
                all_comments.add(comment)
                # all_dates += [date]

            # print(str(len(all_dates)))
            import openpyxl
            wb = openpyxl.load_workbook('shakib_comments.xlsx')
            sheet = wb.active
            global TotalCount
            for c in all_comments:
                TotalCount += 1
                print(str(TotalCount) + ": " + c)
                sheet["A"+str(TotalCount)] = TotalCount
                sheet["B"+str(TotalCount)] = c
            wb.save('shakib_comments.xlsx')
    except Exception as e:
        print(str(e))

def comments():
    line_number = 1
    s1 = ""
    with open('shakib_links.txt', encoding='utf-8') as a_file:
        for a_link in a_file:
            s = a_link.strip()
            # if (s != " "):
            try:
                generate_comments(s)
                line_number += 1
                # print(str(line_number))
            except Exception as e:
                print("Here " + str(e))
    # print("Total Count : ",TotalCount)

def test_generate_comments(theurl):
    thepage = urllib.request.urlopen(theurl).read()
    soup = BeautifulSoup(thepage, "html.parser")

    # target = open("shakib_links.txt", 'a')
    # target.write("Label: " + label + '\n')
    try:
        # line1 = soup.find_all('div', {"class": "teams"}).find('a')
        div = soup.find_all('div', 'user_comment')
        # links = soup.find_all('link', 'all')
        # print(str(len(div)))
        for d in div:
            comments = d.find_all('p')
            dates = d.find_all('span', 'date')
            print(str(len(comments)))
            print(str(len(dates)))
            all_comments = set()
            all_dates = []
            for a_comment,a_date in comments,dates:
                # print(a_comment.text)
                comment = a_comment.text.strip()
                date = a_date.text.strip()
                if comment=="No featured comments at the moment." or comment in all_comments:
                    continue
                all_comments.add(comment)
                all_dates += [date]

            # print(str(len(all_dates)))
            import openpyxl
            # wb = openpyxl.load_workbook('shakib_comments.xlsx')
            # sheet = wb.active
            global TotalCount
            for c,d in all_comments,all_dates:
                TotalCount += 1
                print(str(TotalCount) + ": " + c + "\nDate: " + d)
            #     sheet["A"+str(TotalCount)] = TotalCount
            #     sheet["B"+str(TotalCount)] = c
            # wb.save('shakib_comments.xlsx')
    except Exception as e:
        print(str(e))
    # target.write('\n\n\n')

def main():
    # theurl = "http://www.espncricinfo.com/rss/content/story/feeds/56143.rss"
    # generate_links(theurl)            # generate article links from rss feed
    # comments()                        # extract comments from each link and save in excel
    # test_generate_comments("http://www.espncricinfo.com/bangladesh-premier-league-2015-16/content/story/944017.html?CMP=OTC-RSS")

    import openpyxl
    wb = openpyxl.load_workbook('shakib_comments.xlsx')
    sheet = wb.active
    preprocessing_clause_level.main(sheet.max_row)          # once comments are ready, preprocess them at clause level by running another python script

main()


